from openpyxl import load_workbook
from pathlib import Path
import config  # 引入配置文件
from utils.loger_util import APITestLogger


class ExcelReader:

    def read_excel(self, file_name, sheet_number=None):
        logger = APITestLogger.setup_logger()
        file_path = Path(config.datas_path) / file_name

        try:
            # 检查文件是否存在
            if not file_path.exists():
                logger.error(f"文件 {file_path} 不存在。")
                return ""

            # 加载Excel文件
            workbook = load_workbook(file_path)
            logger.info(f"成功加载文件 {file_path}。")

            # 选择工作表
            if sheet_number is None:
                sheet = workbook.active
                logger.info(f"使用默认工作表 {sheet.title}。")
            else:
                sheet_names = workbook.sheetnames
                if sheet_number > 0 and sheet_number <= len(sheet_names):
                    sheet = workbook[sheet_names[sheet_number - 1]]
                    logger.info(f"使用指定工作表 {sheet.title} (编号 {sheet_number})。")
                else:
                    logger.error(f"工作表编号 {sheet_number} 无效，必须在1和{len(sheet_names)}之间。")
                    return ""

            # 检查工作表是否为空
            if sheet.max_row <= 1:
                logger.warning(f"工作表 {sheet.title} 为空。")
                return ""

            # 提取表头
            headers = [cell.value for cell in sheet[1]]
            logger.debug(f"表头: {headers}")

            # 提取数据行
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 将None替换为""
                row = [cell if cell is not None else "" for cell in row]
                data.append(row)
                logger.debug(f"读取的数据行: {row}")

            if not data:
                logger.warning(f"工作表 {sheet.title} 中没有有效数据行。")
                return ""

            return data
        except Exception as e:
            logger.error(f"读取文件 {file_path} 时出错: {str(e)}")
            return ""


# 使用示例
if __name__ == "__main__":
    reader = ExcelReader()  # 指定文件名称和工作表编号
    data = reader.read_excel("mall登录.xlsx", 1)
    if data != "":
        print("读取的数据:")
        for row in data:
            print(row)
    else:
        print("读取的数据为空。")
